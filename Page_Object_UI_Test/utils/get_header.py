#!/usr/bin/env python3
# encoding=utf-8
# @Time : 4/9/2021 9:43 PM
# @Author : xiaozhan Li
# @Email : lxz_20081025@163.com
# @File : get_excel_data.py
from openpyxl import load_workbook
from ddt import ddt, data,unpack

pre_data = [1, 3]
@ddt
class GetExcelData(object):
    def __init__(self, excel_name, sheet_name):
        self.excel_name = excel_name
        self.sheet_name = sheet_name
        self.wb = load_workbook(self.excel_name)
    @data(pre_data)
    def test_ddt(self, item):
        print(item)

    def get_header(self):
        sheet = self.wb[self.sheet_name]
        header = []
        for j in range(1, sheet.max_column+1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):
        sheet = self.wb[self.sheet_name]
        header = self.get_header()
        test_data = []
        for i in range(2, sheet.max_row + 1):   # test data starts from second row
            sub_data = {}
            for j in range(1, sheet.max_column + 1):
                sub_data[header[j-1]] = sheet.cell(i, j).value
            test_data.append(sub_data)
        return test_data


if __name__ == '__main__':
    pe = GetExcelData('xiaozhan.xlsx', 'test')
    print(f'header is {pe.get_header()}')
    print(f'test data is {pe.get_data()}')



