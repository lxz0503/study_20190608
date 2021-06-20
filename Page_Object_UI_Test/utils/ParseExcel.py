#!/usr/bin/env python3
# coding=utf-8
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# """Operate excel file.For normal operation this script is enough"""
class ParseExcel(object):
    def __init__(self, excel_path_name):
        self.workbook = None
        self.excel_file = excel_path_name
        self.workbook = load_workbook(self.excel_file)
        self.font = Font(color=None)
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def get_sheet_by_name(self, sheet_name):    # get sheet object
        try:
            return self.workbook[sheet_name]
        except Exception as e:
            raise e

    def get_sheet_by_index(self, sheet_index):   # get sheet by index
        try:
            sheet_name = self.workbook.sheetnames[sheet_index]
            return self.workbook[sheet_name]
        except Exception as e:
            raise e

    def get_rows_num(self, sheet):  # sheet needs to be dependent on class instance
        return sheet.max_row

    def get_cols_num(self, sheet):
        return sheet.max_column

    def get_start_row_num(self, sheet):
        return sheet.min_row

    def get_start_col_num(self, sheet):
        return sheet.min_column

    def get_row_values(self, sheet):
        # get values in all rows, a tuple,  this returns a generator
        try:
            return sheet.rows        # it is a generator
        except Exception as e:
            raise e

    def get_each_row_values(self, sheet, row):    # this can get values in specific row
        columns = self.get_cols_num(sheet)
        row_data = []
        for i in range(1, columns + 1):
            cell_value = sheet.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def get_col_values(self, sheet):
        # get values in all columns, a tuple
        try:
            return sheet.columns     # it is a generator for all columns
        except Exception as e:
            raise e

    def get_col_values_by_letter(self, sheet, col):    # just get one column like sheet['C']
        try:
            values = sheet[col]
            return [x.value for x in values if x.value is not None]    # return a list, you can define any format as you like
        except Exception as e:
            raise e

    def get_col_values_by_index(self, sheet, col):   # get value in specific column
        rows = self.get_rows_num(sheet)
        column_data = []
        for i in range(1, rows + 1):
            cell_value = sheet.cell(row=i, column=col).value
            column_data.append(cell_value)
        return column_data

    def get_cell_value(self, sheet, row_num, col_num):         # if you just want to get one cell value
        try:
            return sheet.cell(row=row_num, column=col_num).value
        except Exception as e:
            raise e

    def write_excel(self, sheet, row_num, col_num, content):    # just write one cell
        try:
            sheet.cell(row=row_num, column=col_num).value = content
            self.workbook.save(self.excel_file)
        except Exception as e:
            raise e


if __name__ == '__main__':
    pe = ParseExcel('xiaozhan.xlsx')
    # sheet = pe.get_sheet_by_index(0)    # get sheet object, this is important for later procedures
    sheet = pe.get_sheet_by_name('test')  # test is the sheet name
    print(sheet.title)    # the first sheet name
    # print(pe.get_cols_num(sheet))
    # get values in all rows and print each row
    r = pe.get_row_values(sheet)
    for row in r:
        for cell in row:
            print(cell.value, end=' ')    # # you can also put them into a list using append as you like
        print()    # this will help to split each row

    # get values in all rows and print each row
    print('print value by column====================== as below')
    r = pe.get_col_values(sheet)
    for col in r:
        for cell in col:
            print(cell.value, end=' ')    # you can also put them into a list as you like
        print()  # this will help to split each row

    # get values in specific line
    print(pe.get_each_row_values(sheet, 2))         # [1000, 2000, 3000, None]
    # get value in specific cell
    print(pe.get_cell_value(sheet, 3, 3))
    # write value into specific cell
    pe.write_excel(sheet, 4, 4, 'xiaozhan')
    # get values in column C
    print(pe.get_col_values_by_letter(sheet, 'C'))   # ['tianjin', 3000, 'dongcheng', None]

    # get values in column 1
    print(pe.get_col_values_by_index(sheet, 1))   # ['beijing', 1000, 'shunyi', None]




