#!/usr/bin/env python3
# coding=utf-8

from openpyxl import load_workbook, Workbook
import glob


def get_data_from_excel(path, new_excel):
    """一个目录下所有的excel文件，这些文件都有相同的表头。选取某一列数据，写到一个新的excel"""
    # record data into a new excel
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    flag = 0     # 确认新表是否已经添加表头

    for file in glob.glob(path + '/' + 'bug*.xlsx'):
        workbook = load_workbook(path + '/' + 'bug_copy.xlsx')  # 打开已经存在的excel
        sheet = workbook.active

        # select rows
        all_count = sheet['B']   # 选择B列
        row_lst = []
        for cell in all_count:
            if isinstance(cell.value, int) and cell.value > 20:
                print(cell.row)    # 显示符合条件的行号
                row_lst.append(cell.row)   # 把行号存储在列表

        # create header(the first row)
        if not flag:
            header = sheet[1]       # the first row
            header_lst = []
            for cell in header:
                header_lst.append(cell.value)
            new_sheet.append(header_lst)
            flag = 1

        # 从旧表中根据行号提取符合条件的行，并遍历单元格，以列表形式写入新表
        for row in row_lst:
            data_lst = []
            for cell in sheet[row]:   # 遍历每一行
                data_lst.append(cell.value)
            new_sheet.append(data_lst)
    # generate new excel
    new_workbook.save(path + '/' + new_excel)


if __name__ == '__main__':
    path = 'excel'
    new_excel = 'new.xlsx'
    get_data_from_excel(path, new_excel)