#!/usr/bin/env python3
# coding=utf-8

from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('example.xlsx')
# ws = wb.worksheets
# # print(ws)
print(wb.get_sheet_names())     # this is a list
#
ws = wb.get_sheet_by_name('student')
print('type is {0}, values are {1}'.format(type(ws.values), ws.values))

# for i in ws.values:
#     print(type(i), i)

for row in ws.rows:
    print(*[cell.value for cell in row])    # ????不明白这个语法

# create a new workbook  and write data
wb = Workbook()
ws = wb.active
print(ws)      # default is sheet
ws.title = 'student'
print(ws)        # now it is student
wb.create_sheet(index=0, title='new sheet')     # create a new sheet
ws = wb.active
# ws = wb.get_sheet_names()
print(ws)
ws['A1'] = 'hello'
ws['A2'] = 'beijing'
# wb.save('python-openpyxl.xlsx')    # save data into excel
