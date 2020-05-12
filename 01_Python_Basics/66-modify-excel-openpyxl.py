#!/usr/bin/env python3
# coding=utf-8

import openpyxl
from openpyxl import Workbook
import shutil

def process_worksheet(sheet):
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2

    for row in sheet.iter_rows(min_row=2, min_col=3):    # start from row=2 and column=3
        # print(row, row[0].row)   # row[0].row can locate the row number
        scores = [cell.value for cell in row]   # this is a list
        sum_score = sum(scores)
        avg_score = sum_score / len(scores)

    # calculate avg and store them into the last two columns
        sheet.cell(row=row[0].row, column=avg_column).value = avg_score
        sheet.cell(row=row[0].row, column=sum_column).value = sum_score

    # set column title
    sheet.cell(row=1, column=avg_column).value = 'avg'
    sheet.cell(row=1, column=sum_column).value = 'sum'

def main():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('student')
    process_worksheet(sheet)
    wb.save('example_copy.xlsx')
    # shutil.move('example_copy.xlsx', 'example.xlsx')   # use new file to replace old one

if __name__ == '__main__':
    main()

# print(row, row[0].row)
# (<Cell 'student'.C2>, <Cell 'student'.D2>, <Cell 'student'.E2>) 2
# (<Cell 'student'.C3>, <Cell 'student'.D3>, <Cell 'student'.E3>) 3
# (<Cell 'student'.C4>, <Cell 'student'.D4>, <Cell 'student'.E4>) 4
# (<Cell 'student'.C5>, <Cell 'student'.D5>, <Cell 'student'.E5>) 5
# (<Cell 'student'.C6>, <Cell 'student'.D6>, <Cell 'student'.E6>) 6